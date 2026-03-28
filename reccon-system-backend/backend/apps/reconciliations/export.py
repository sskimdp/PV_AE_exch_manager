from __future__ import annotations

from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from apps.messages.models import Message


DATE_TIME_FORMAT = "dd.mm.yyyy hh:mm"
META_LABEL_FONT = Font(bold=True)
TITLE_FONT = Font(size=14, bold=True)
TABLE_HEADER_FONT = Font(bold=True)
TABLE_HEADER_FILL = PatternFill(fill_type="solid", fgColor="D9E2F3")


def _safe_subject(value: str | None) -> str:
    text = str(value or "").strip()
    return text or "Без темы"


def _map_status(value: str | None) -> str:
    normalized = str(value or "").strip().lower()
    if normalized == Message.STATUS_CONFIRMED:
        return "Подтверждено"
    if normalized == Message.STATUS_READ:
        return "Прочитано"
    if normalized == Message.STATUS_PENDING:
        return "Ожидает подтверждения"
    if normalized == Message.STATUS_DRAFT:
        return "Черновик"
    return str(value or "Ожидает подтверждения").strip() or "Ожидает подтверждения"


def _to_excel_datetime(value):
    if not value:
        return None

    if timezone.is_aware(value):
        value = timezone.localtime(value)

    return value.replace(tzinfo=None) if getattr(value, "tzinfo", None) else value


def _period_label(reconciliation) -> str:
    return (
        f"{reconciliation.period_start.strftime('%d.%m.%Y')} - "
        f"{reconciliation.period_end.strftime('%d.%m.%Y')}"
    )


def _exported_at_label() -> str:
    return timezone.localtime().strftime("%d.%m.%Y %H:%M")


def _iter_stage_items(stage):
    return sorted(stage.items.all(), key=lambda item: item.id)


def _stage_row(item, *, include_stage: bool = False):
    message = getattr(item, "message", None)
    snapshot_status = item.status_snapshot or getattr(message, "status", "") or ""

    outgoing_number = getattr(message, "sender_number", "") or ""
    incoming_number = (
        getattr(message, "receiver_number", "") or ""
        if snapshot_status == Message.STATUS_CONFIRMED
        else ""
    )
    number = incoming_number if snapshot_status == Message.STATUS_CONFIRMED else outgoing_number

    row = []
    if include_stage:
        row.append(item.stage.stage_number)

    row.extend(
        [
            number,
            outgoing_number,
            incoming_number,
            _safe_subject(item.subject_snapshot or getattr(message, "subject", "")),
            _to_excel_datetime(item.sent_at_snapshot or getattr(message, "created_at", None)),
            _to_excel_datetime(item.confirmed_at_snapshot),
            _map_status(snapshot_status),
        ]
    )
    return row


def _write_sheet_header(worksheet, *, reconciliation, sheet_label: str, total_columns: int):
    worksheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=total_columns)
    title_cell = worksheet.cell(row=1, column=1, value="Сверка сообщений")
    title_cell.font = TITLE_FONT
    title_cell.alignment = Alignment(horizontal="left", vertical="center")

    metadata_rows = [
        ("Master-компания", reconciliation.master_company.name),
        ("Slave-компания", reconciliation.slave_company.name),
        ("Период", _period_label(reconciliation)),
        ("Лист", sheet_label),
        ("Дата экспорта", _exported_at_label()),
    ]

    row_index = 2
    for label, value in metadata_rows:
        label_cell = worksheet.cell(row=row_index, column=1, value=label)
        value_cell = worksheet.cell(row=row_index, column=2, value=value)
        label_cell.font = META_LABEL_FONT
        value_cell.alignment = Alignment(horizontal="left", vertical="center")
        row_index += 1

    worksheet.row_dimensions[1].height = 24
    worksheet.freeze_panes = "A8"


def _apply_table_formatting(worksheet, *, has_stage_column: bool):
    widths = {
        "A": 12 if has_stage_column else 18,
        "B": 18,
        "C": 18,
        "D": 18,
        "E": 44 if has_stage_column else 44,
        "F": 22 if has_stage_column else 22,
        "G": 22 if has_stage_column else 22,
        "H": 24,
    }

    for column_letter, width in widths.items():
        if not has_stage_column and column_letter == "H":
            continue
        worksheet.column_dimensions[column_letter].width = width

    text_columns = ["B", "C", "D"] if has_stage_column else ["A", "B", "C"]
    datetime_columns = ["F", "G"] if has_stage_column else ["E", "F"]

    for cell in worksheet[8]:
        cell.font = TABLE_HEADER_FONT
        cell.fill = TABLE_HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in worksheet.iter_rows(min_row=9, max_row=worksheet.max_row):
        for column_letter in text_columns:
            row[ord(column_letter) - ord("A")].number_format = "@"

        for column_letter in datetime_columns:
            cell = row[ord(column_letter) - ord("A")]
            if cell.value:
                cell.number_format = DATE_TIME_FORMAT

        subject_column_letter = "E" if has_stage_column else "D"
        status_column_letter = "H" if has_stage_column else "G"

        row[ord(subject_column_letter) - ord("A")].alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True,
        )
        row[ord(status_column_letter) - ord("A")].alignment = Alignment(
            horizontal="left",
            vertical="center",
        )

    worksheet.auto_filter.ref = (
        f"A8:{get_column_letter(worksheet.max_column)}{max(worksheet.max_row, 8)}"
    )


def _append_stage_sheet(workbook, *, reconciliation, stage):
    worksheet = workbook.create_sheet(title=f"Этап {stage.stage_number}")
    headers = [
        "Номер",
        "Исходящий номер",
        "Входящий номер",
        "Тема",
        "Дата отправки",
        "Дата подтверждения",
        "Статус",
    ]

    _write_sheet_header(
        worksheet,
        reconciliation=reconciliation,
        sheet_label=f"Этап {stage.stage_number}",
        total_columns=len(headers),
    )
    worksheet.append([])
    worksheet.append(headers)

    for item in _iter_stage_items(stage):
        worksheet.append(_stage_row(item))

    _apply_table_formatting(worksheet, has_stage_column=False)


def _append_all_stages_sheet(workbook, *, reconciliation, stages):
    worksheet = workbook.create_sheet(title="Все этапы")
    headers = [
        "Этап",
        "Номер",
        "Исходящий номер",
        "Входящий номер",
        "Тема",
        "Дата отправки",
        "Дата подтверждения",
        "Статус",
    ]

    _write_sheet_header(
        worksheet,
        reconciliation=reconciliation,
        sheet_label="Все этапы",
        total_columns=len(headers),
    )
    worksheet.append([])
    worksheet.append(headers)

    for stage in stages:
        for item in _iter_stage_items(stage):
            worksheet.append(_stage_row(item, include_stage=True))

    _apply_table_formatting(worksheet, has_stage_column=True)


def build_reconciliation_export_workbook(*, reconciliation, scope: str, stage_number: int | None = None):
    workbook = Workbook()
    workbook.remove(workbook.active)

    stages = sorted(reconciliation.stages.all(), key=lambda stage: stage.stage_number)

    if scope == "all":
        _append_all_stages_sheet(
            workbook,
            reconciliation=reconciliation,
            stages=stages,
        )
        for stage in stages:
            _append_stage_sheet(
                workbook,
                reconciliation=reconciliation,
                stage=stage,
            )
        return workbook

    target_stage = next(
        (stage for stage in stages if int(stage.stage_number) == int(stage_number)),
        None,
    )
    if target_stage is None:
        raise ValueError("Stage not found for this reconciliation.")

    _append_stage_sheet(
        workbook,
        reconciliation=reconciliation,
        stage=target_stage,
    )
    return workbook
